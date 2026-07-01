"""
Leitura de planilhas de itens. Detecta as colunas de descrição, NCM e UF pelo
cabeçalho (linha 1), sem alterar nada. Funciona sobre um worksheet openpyxl —
o carregamento do arquivo é separado (load_items) para permitir teste em memória.

⚠️ Escala (plano §3): planilhas reais de 40k linhas com fórmulas são pesadas em
memória. Preferir read_only=True na leitura; impor limite de tamanho por request.
"""
from __future__ import annotations

from typing import Any

from services.shared.contracts.fiscal import normalize_descricao

# Aliases de cabeçalho (normalizados) → campo lógico.
_DESC_ALIASES = {"descricao", "descricao do produto", "produto", "item", "mercadoria", "sku"}
_NCM_ALIASES = {"ncm", "ncm codigo", "codigo ncm", "classificacao fiscal"}
_UF_ALIASES = {"uf", "uf destino", "estado", "uf de destino", "destino"}


def detect_columns(headers: list[Any]) -> dict[str, int | None]:
    """
    Mapeia campos lógicos → índice de coluna (1-based) a partir do cabeçalho.
    Retorna {'descricao': int|None, 'ncm': int|None, 'uf': int|None}.
    """
    colmap: dict[str, int | None] = {"descricao": None, "ncm": None, "uf": None}
    for idx, raw in enumerate(headers, start=1):
        if raw is None:
            continue
        norm = normalize_descricao(str(raw))
        if colmap["descricao"] is None and norm in _DESC_ALIASES:
            colmap["descricao"] = idx
        elif colmap["ncm"] is None and norm in _NCM_ALIASES:
            colmap["ncm"] = idx
        elif colmap["uf"] is None and norm in _UF_ALIASES:
            colmap["uf"] = idx
    return colmap


def _clean_ncm(value: Any) -> str | None:
    """Extrai 8 dígitos de um valor de NCM (pode vir com pontos/espaços)."""
    if value is None:
        return None
    digitos = "".join(ch for ch in str(value) if ch.isdigit())
    return digitos.zfill(8) if 0 < len(digitos) <= 8 else None


def read_rows(ws: Any, colmap: dict[str, int | None]) -> list[dict[str, Any]]:
    """
    Lê as linhas de dados (a partir da linha 2) usando o mapa de colunas.
    Retorna dicts {'row': n, 'descricao', 'ncm_hint', 'uf'}.
    """
    def _cell(row: tuple, field: str) -> Any:
        col = colmap.get(field)
        return row[col - 1] if col and col - 1 < len(row) else None

    rows: list[dict[str, Any]] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        descricao = _cell(row, "descricao")
        uf = _cell(row, "uf")
        rows.append({
            "row": r_idx,
            "descricao": str(descricao).strip() if descricao is not None else "",
            "ncm_hint": _clean_ncm(_cell(row, "ncm")),
            "uf": str(uf).strip().upper() if uf is not None else None,
        })
    return rows


def load_items(path: str) -> tuple[dict[str, int | None], list[dict[str, Any]]]:
    """Abre a planilha em modo somente-leitura e extrai (colmap, linhas)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        colmap = detect_columns(list(headers))
        rows = read_rows(ws, colmap)
    finally:
        wb.close()
    return colmap, rows
