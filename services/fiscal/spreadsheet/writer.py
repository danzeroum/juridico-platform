"""
Escrita do enriquecimento na planilha SEM quebrar fórmulas existentes.

Estratégia (plano §4.4 da discussão): adicionar as colunas novas À DIREITA da
planilha original, escrevendo apenas células novas. Nunca reescrever/mover células
existentes — openpyxl preserva as fórmulas originais (carregadas como texto) ao
salvar. Cada `NcmTriageResult` é escrito na linha correspondente ao item.

⚠️ Ressalva (plano §3): openpyxl pode não preservar gráficos/tabelas dinâmicas de
planilhas complexas e é pesado em memória para 40k linhas. Testar com planilhas
reais dos clientes e medir memória por job.
"""
from __future__ import annotations

from typing import Any

from services.shared.contracts.fiscal import NcmTriageResult

# Colunas de enriquecimento acrescentadas à direita, em ordem.
ENRICHMENT_HEADERS = [
    "NCM sugerido",
    "Confiança NCM",
    "Fonte da regra",
    "ICMS interno efetivo (%)",
    "ICMS interestadual (%)",
    "DIFAL (%)",
    "Categoria",
    "Conflito?",
    "Observações",
]


def _row_values(result: NcmTriageResult) -> list[Any]:
    ncm = result.suggested_ncm
    icms = result.icms
    return [
        ncm.ncm_codigo if ncm else None,
        ncm.confidence if ncm else None,
        ncm.fonte_regra.value if ncm else None,
        icms.interna_efetiva_pct,
        icms.interestadual_pct,
        icms.difal_pct,
        result.categoria,
        "SIM" if result.conflito_detectado else "NÃO",
        " | ".join(result.observacoes) if result.observacoes else None,
    ]


def append_enrichment(
    ws: Any,
    results_by_row: dict[int, NcmTriageResult],
    *,
    header_row: int = 1,
) -> int:
    """
    Acrescenta as colunas de enriquecimento à direita da última coluna existente.
    `results_by_row` mapeia número da linha (1-based) → resultado da triagem.
    Retorna a primeira coluna nova utilizada.
    """
    start_col = ws.max_column + 1

    # Cabeçalhos das colunas novas.
    for offset, title in enumerate(ENRICHMENT_HEADERS):
        ws.cell(row=header_row, column=start_col + offset, value=title)

    # Valores por linha (só onde houve triagem; células existentes ficam intactas).
    for row_idx, result in results_by_row.items():
        for offset, value in enumerate(_row_values(result)):
            ws.cell(row=row_idx, column=start_col + offset, value=value)

    return start_col


def enrich_workbook(in_path: str, out_path: str, results_by_row: dict[int, NcmTriageResult]) -> None:
    """Carrega a planilha (preservando fórmulas), acrescenta colunas e salva em out_path."""
    from openpyxl import load_workbook

    wb = load_workbook(in_path)  # data_only=False (default) → mantém fórmulas
    try:
        append_enrichment(wb.active, results_by_row)
        wb.save(out_path)
    finally:
        wb.close()
