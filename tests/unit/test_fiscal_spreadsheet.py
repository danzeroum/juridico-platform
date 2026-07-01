"""
Testes de leitura/escrita de planilhas. Asserção CRÍTICA: o writer NÃO quebra
fórmulas existentes ao acrescentar colunas à direita.
"""
from __future__ import annotations

from openpyxl import Workbook, load_workbook

from services.fiscal.spreadsheet.reader import (
    detect_columns,
    load_items,
    read_rows,
)
from services.fiscal.spreadsheet.writer import (
    ENRICHMENT_HEADERS,
    append_enrichment,
    enrich_workbook,
)
from services.shared.contracts.fiscal import (
    FonteRegra,
    IcmsResolution,
    NcmCandidate,
    NcmTriageResult,
)


def _make_wb_with_formula():
    wb = Workbook()
    ws = wb.active
    ws.append(["Descrição", "NCM", "UF", "Preço", "Total"])
    ws.append(["Notebook Dell", "8471.30.12", "RJ", 5000, "=D2*1.1"])
    ws.append(["Água mineral", "22021000", "SP", 3, "=D3*1.1"])
    return wb, ws


def _result(desc, ncm):
    return NcmTriageResult(
        sku_descricao=desc,
        suggested_ncm=NcmCandidate(ncm_codigo=ncm, descricao=desc, confidence=0.9, fonte_regra=FonteRegra.TIPI),
        icms=IcmsResolution(interna_efetiva_pct=20.0, interestadual_pct=12.0, difal_pct=8.0),
        conflito_detectado=False,
        observacoes=["ok"],
    )


class TestReader:
    def test_detect_columns(self):
        colmap = detect_columns(["Descrição", "NCM", "UF", "Preço"])
        assert colmap == {"descricao": 1, "ncm": 2, "uf": 3}

    def test_read_rows_limpa_ncm_e_uf(self):
        _wb, ws = _make_wb_with_formula()
        colmap = detect_columns(["Descrição", "NCM", "UF", "Preço", "Total"])
        rows = read_rows(ws, colmap)
        assert rows[0]["ncm_hint"] == "84713012"  # pontos removidos, pad 8
        assert rows[0]["uf"] == "RJ"
        assert rows[0]["descricao"] == "Notebook Dell"

    def test_load_items_de_arquivo(self, tmp_path):
        wb, _ws = _make_wb_with_formula()
        p = tmp_path / "in.xlsx"
        wb.save(p)
        colmap, rows = load_items(str(p))
        assert colmap["descricao"] == 1
        assert len(rows) == 2

    def test_load_items_from_bytes(self):
        # Caminho usado pelo worker: lê a planilha dos bytes vindos do MinIO.
        from io import BytesIO

        from services.fiscal.spreadsheet.reader import load_items_from_bytes

        wb, _ws = _make_wb_with_formula()
        buf = BytesIO()
        wb.save(buf)
        colmap, rows = load_items_from_bytes(buf.getvalue())
        assert colmap["descricao"] == 1
        assert len(rows) == 2
        assert rows[0]["ncm_hint"] == "84713012"


class TestWriter:
    def test_append_nao_altera_formula(self):
        wb, ws = _make_wb_with_formula()
        start = append_enrichment(ws, {2: _result("Notebook Dell", "84713012")})
        # Fórmula original intacta.
        assert ws["E2"].value == "=D2*1.1"
        # Cabeçalhos novos à direita.
        assert ws.cell(row=1, column=start).value == ENRICHMENT_HEADERS[0]
        # Valores escritos na linha do item.
        assert ws.cell(row=2, column=start).value == "84713012"

    def test_enrich_workbook_preserva_formula_no_disco(self, tmp_path):
        wb, _ws = _make_wb_with_formula()
        in_p = tmp_path / "in.xlsx"
        out_p = tmp_path / "out.xlsx"
        wb.save(in_p)
        enrich_workbook(str(in_p), str(out_p), {2: _result("Notebook Dell", "84713012")})

        reloaded = load_workbook(out_p)  # mantém fórmulas
        ws2 = reloaded.active
        assert ws2["E2"].value == "=D2*1.1"
        assert ws2.cell(row=1, column=ws2.max_column - len(ENRICHMENT_HEADERS) + 1).value == ENRICHMENT_HEADERS[0]
        reloaded.close()
